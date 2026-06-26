from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import extract
from app.database import get_db
from app.models import ReturnMaterial, ReturnReason
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime

router = APIRouter()


@router.get("/excel")
async def export_excel(
    status: str = None,
    month: int = None,
    year: int = None,
    db: Session = Depends(get_db),
):
    query = db.query(ReturnMaterial).join(ReturnReason)
    if status:
        query = query.filter(ReturnMaterial.status == status)
    if month:
        query = query.filter(extract("month", ReturnMaterial.created_at) == month)
    if year:
        query = query.filter(extract("year", ReturnMaterial.created_at) == year)

    returns_list = query.order_by(ReturnMaterial.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Return Materials"

    # Header style
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    headers = ["No", "Lot Ref", "Qty", "Reason", "Condition", "Status", "Destination", "Note", "Created At"]
    col_widths = [6, 15, 12, 20, 15, 12, 12, 30, 20]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[chr(64 + col)].width = width

    # Data rows
    for idx, ret in enumerate(returns_list, 1):
        row = idx + 1
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=ret.lot_ref)
        ws.cell(row=row, column=3, value=float(ret.qty))
        ws.cell(row=row, column=4, value=ret.reason.name if ret.reason else "")
        ws.cell(row=row, column=5, value=ret.condition or "")
        ws.cell(row=row, column=6, value=ret.status)
        ws.cell(row=row, column=7, value=ret.destination or "")
        ws.cell(row=row, column=8, value=ret.note or "")
        ws.cell(row=row, column=9, value=ret.created_at.strftime("%Y-%m-%d %H:%M") if ret.created_at else "")

    # Summary row
    summary_row = len(returns_list) + 3
    ws.cell(row=summary_row, column=1, value="Total Returns:")
    ws.cell(row=summary_row, column=2, value=len(returns_list))
    ws.cell(row=summary_row + 1, column=1, value="Export Date:")
    ws.cell(row=summary_row + 1, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M"))

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"return_materials_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
