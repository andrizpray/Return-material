from fastapi import APIRouter, Request, Depends, Form, HTTPException, UploadFile, File
from fastapi.responses import RedirectResponse, StreamingResponse, Response, JSONResponse
from sqlalchemy.orm import Session
from app.database import get_db
from app.models import (
    ReturnMaterial, BeritaAcara, BeritaAcaraItem, BeritaAcaraPhoto,
)
from fastapi.templating import Jinja2Templates
from app.templates import TEMPLATE_DIR
from datetime import datetime
from openpyxl import load_workbook
import io, os, uuid

router = APIRouter()
templates = Jinja2Templates(directory=TEMPLATE_DIR)

BA_UPLOAD_DIR = os.path.join(
    os.path.dirname(os.path.dirname(__file__)), "uploads", "ba"
)
os.makedirs(BA_UPLOAD_DIR, exist_ok=True)

TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    "templates",
    "berita_acara_template.xlsx",
)


def _indonesian_day(dt: datetime) -> str:
    days = {
        0: "Senin", 1: "Selasa", 2: "Rabu", 3: "Kamis",
        4: "Jumat", 5: "Sabtu", 6: "Minggu",
    }
    return days[dt.weekday()]


def _format_date(dt: datetime) -> str:
    return dt.strftime("%d/%m/%Y")


def _generate_nomor_bap(dt: datetime) -> str:
    """Auto-generate nomor BAP: BAP/RTN/DD/MM/ROMAN."""
    months = [
        "I", "II", "III", "IV", "V", "VI",
        "VII", "VIII", "IX", "X", "XI", "XII",
    ]
    day = dt.strftime("%d")
    month = months[dt.month - 1]
    year_short = dt.strftime("%y")
    return f"BAP/RTN/{day}/{month}/{year_short}"


def _render_list_fragment(db: Session) -> str:
    """Render berita acara list as HTML fragment (no base.html wrapper)."""
    items = (
        db.query(BeritaAcara)
        .order_by(BeritaAcara.created_at.desc())
        .all()
    )
    tpl = templates.get_template("pages/berita_acara_list.html")
    return tpl.render({"active_page": "berita_acara", "items": items})


@router.get("/")
async def list_berita_acara(request: Request, db: Session = Depends(get_db)):
    items = (
        db.query(BeritaAcara)
        .order_by(BeritaAcara.created_at.desc())
        .all()
    )
    return templates.TemplateResponse(
        request,
        "pages/berita_acara_list.html",
        {"active_page": "berita_acara", "items": items},
    )


@router.get("/create")
async def create_form(
    request: Request,
    return_ids: str = "",
    db: Session = Depends(get_db),
):
    now = datetime.now()
    return templates.TemplateResponse(
        request,
        "pages/berita_acara_form.html",
        {
            "active_page": "berita_acara",
            "now": now,
            "default_nomor": _generate_nomor_bap(now),
        },
    )


@router.post("/create")
async def create_berita_acara(
    request: Request,
    # Section I
    nomor_bap: str = Form(...),
    tanggal: str = Form(...),
    hal: str = Form("Broke / Return / Damage Product / Out Spec / Etc"),
    ditujukan_kepada: str = Form("QC"),
    cc: str = Form(""),
    # SJ info
    no_sj: str = Form(""),
    tanggal_sj: str = Form(""),
    no_sj_return: str = Form(""),
    customer_name: str = Form(...),
    customer_address: str = Form(""),
    # Signatures
    admin_delivery: str = Form(""),
    karu_delivery: str = Form(""),
    kabag_scm: str = Form(""),
    sopir_nopol: str = Form(""),
    # Items (multi-value) — manual input
    lot_ids: list[str] = Form([]),
    qtys: list[str] = Form([]),
    jenis_barangs: list[str] = Form([]),
    rew_ids: list[str] = Form([]),
    keterangans: list[str] = Form([]),
    db: Session = Depends(get_db),
):
    dt = datetime.strptime(tanggal, "%Y-%m-%d")
    tgl_sj = datetime.strptime(tanggal_sj, "%Y-%m-%d") if tanggal_sj else None

    ba = BeritaAcara(
        nomor_bap=nomor_bap,
        tanggal=dt,
        hal=hal,
        ditujukan_kepada=ditujukan_kepada,
        cc=cc or None,
        no_sj=no_sj or None,
        tanggal_sj=tgl_sj,
        no_sj_return=no_sj_return or None,
        customer_name=customer_name,
        customer_address=customer_address or None,
        admin_delivery=admin_delivery or None,
        karu_delivery=karu_delivery or None,
        kabag_scm=kabag_scm or None,
        sopir_nopol=sopir_nopol or None,
    )
    db.add(ba)
    db.flush()

    for i in range(len(lot_ids)):
        lot = lot_ids[i].strip() if i < len(lot_ids) else ""
        qty_str = qtys[i].strip() if i < len(qtys) else ""
        if not lot and not qty_str:
            continue  # skip empty rows

        qty_val = None
        if qty_str:
            try:
                qty_val = float(qty_str)
            except ValueError:
                pass

        item = BeritaAcaraItem(
            berita_acara_id=ba.id,
            lot_id=lot or None,
            qty=qty_val,
            jenis_barang=jenis_barangs[i] if i < len(jenis_barangs) and jenis_barangs[i].strip() else None,
            rew_id=rew_ids[i] if i < len(rew_ids) and rew_ids[i].strip() else None,
            keterangan=keterangans[i] if i < len(keterangans) and keterangans[i].strip() else None,
        )
        db.add(item)

    db.commit()
    return RedirectResponse(f"/berita-acara/{ba.id}", status_code=303)


@router.post("/{ba_id}/photo/upload")
async def upload_photo(ba_id: int, photo: UploadFile = File(...), db: Session = Depends(get_db)):
    ba = db.query(BeritaAcara).filter(BeritaAcara.id == ba_id).first()
    if not ba:
        raise HTTPException(404, "Berita Acara not found")

    ext = os.path.splitext(photo.filename)[1].lower()
    if ext not in (".jpg", ".jpeg", ".png", ".gif", ".webp"):
        raise HTTPException(400, "File harus gambar (jpg/png/gif/webp)")

    filename = f"{uuid.uuid4().hex}{ext}"
    filepath = os.path.join(BA_UPLOAD_DIR, filename)
    content = await photo.read()
    with open(filepath, "wb") as f:
        f.write(content)

    bp = BeritaAcaraPhoto(
        berita_acara_id=ba_id,
        filename=filename,
        original_name=photo.filename,
    )
    db.add(bp)
    db.commit()
    db.refresh(bp)

    return JSONResponse({
        "id": bp.id,
        "url": f"/uploads/ba/{filename}",
        "name": photo.filename,
    })


@router.post("/{ba_id}/photo/{photo_id}/delete")
async def delete_photo(ba_id: int, photo_id: int, db: Session = Depends(get_db)):
    photo = db.query(BeritaAcaraPhoto).filter(
        BeritaAcaraPhoto.id == photo_id,
        BeritaAcaraPhoto.berita_acara_id == ba_id,
    ).first()
    if photo:
        filepath = os.path.join(BA_UPLOAD_DIR, photo.filename)
        if os.path.exists(filepath):
            os.remove(filepath)
        db.delete(photo)
        db.commit()
    return Response(content="", media_type="text/html")


@router.get("/{ba_id}")
async def detail_berita_acara(request: Request, ba_id: int, db: Session = Depends(get_db)):
    ba = db.query(BeritaAcara).filter(BeritaAcara.id == ba_id).first()
    if not ba:
        return RedirectResponse("/berita-acara/", status_code=303)
    return templates.TemplateResponse(
        request,
        "pages/berita_acara_detail.html",
        {"active_page": "berita_acara", "ba": ba},
    )


@router.get("/{ba_id}/download")
async def download_excel(ba_id: int, db: Session = Depends(get_db)):
    ba = db.query(BeritaAcara).filter(BeritaAcara.id == ba_id).first()
    if not ba:
        raise HTTPException(404, "Berita Acara not found")

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["DOCK"]

    # === Clear stale template data (Section II - rows 14-28) ===
    merges_to_remove = [mc for mc in list(ws.merged_cells.ranges)
                        if mc.min_row >= 14 and mc.max_row <= 28]
    for mc in merges_to_remove:
        ws.unmerge_cells(str(mc))
    for row in range(14, 29):
        for col in range(1, 13):
            ws.cell(row=row, column=col).value = None

    # === Section I ===
    ws["C53"] = ba.nomor_bap
    ws["C54"] = f"{_indonesian_day(ba.tanggal)}, {_format_date(ba.tanggal)}"
    ws["C55"] = ba.hal or ""
    ws["C56"] = ba.ditujukan_kepada or ""
    ws["C57"] = ba.cc or "-"

    # === SJ Info ===
    sj_text = ""
    if ba.no_sj:
        sj_text = ba.no_sj
        if ba.tanggal_sj:
            sj_text += f" / {_format_date(ba.tanggal_sj)}"
    ws["C60"] = sj_text
    ws["C61"] = ba.no_sj_return or "-"
    ws["C62"] = ba.customer_name
    ws["C63"] = ba.customer_address or ""

    # === Section III: Items ===
    start_row = 69
    total_qty = 0
    total_rolls = 0
    for i, item in enumerate(ba.items):
        row = start_row + i
        ws.cell(row=row, column=1, value=i + 1)  # No
        ws.cell(row=row, column=2, value=item.jenis_barang or "")  # Jenis Barang
        # D = Rew ID, E = Lot ID
        ws.cell(row=row, column=4, value=item.rew_id or "")
        # Use item.lot_id first, fallback to linked return_material
        lot_id = item.lot_id or ""
        if not lot_id and item.return_material:
            lot_id = item.return_material.lot_ref or ""
        ws.cell(row=row, column=5, value=lot_id)
        # F = Qty Kg — use item.qty first, fallback to linked return_material
        qty = 0
        if item.qty:
            qty = float(item.qty)
        elif item.return_material and item.return_material.qty:
            qty = float(item.return_material.qty)
        ws.cell(row=row, column=6, value=qty if qty else "")
        total_qty += qty
        total_rolls += 1
        # G = Keterangan
        keterangan = item.keterangan or ""
        if not keterangan and item.return_material:
            keterangan = item.return_material.condition or item.return_material.note or ""
        ws.cell(row=row, column=7, value=keterangan)

    # TOTAL row
    if ba.items:
        total_row = start_row + len(ba.items)
        ws.cell(row=total_row, column=2, value="TOTAL")
        ws.cell(row=total_row, column=3, value=f"{total_rolls} ROLL{'S' if total_rolls != 1 else ''}")
        ws.cell(row=total_row, column=6, value=total_qty if total_qty else "")

    # === Section IV: Signatures ===
    if ba.admin_delivery:
        ws["A81"] = ba.admin_delivery
    if ba.karu_delivery:
        ws["D81"] = ba.karu_delivery
    if ba.kabag_scm:
        ws["E81"] = ba.kabag_scm
    if ba.sopir_nopol:
        ws["G81"] = ba.sopir_nopol

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Berita_Acara_{ba.nomor_bap.replace('/', '-')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/{ba_id}/delete")
async def delete_berita_acara(ba_id: int, db: Session = Depends(get_db)):
    ba = db.query(BeritaAcara).filter(BeritaAcara.id == ba_id).first()
    if ba:
        db.delete(ba)
        db.commit()
    # Return list fragment for HTMX swap (no base.html wrapper)
    html = _render_list_fragment(db)
    return Response(content=html, media_type="text/html",
                    headers={"HX-Redirect": "/berita-acara/"})
